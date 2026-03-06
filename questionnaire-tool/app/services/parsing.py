import csv
import io
import re
from pathlib import Path
from typing import List

from openpyxl import load_workbook
from pypdf import PdfReader


QUESTION_HINTS = {"question", "questions", "prompt", "item", "control"}


def normalize_text(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def parse_pdf_text(file_bytes: bytes) -> str:
    reader = PdfReader(io.BytesIO(file_bytes))
    pages = [page.extract_text() or "" for page in reader.pages]
    return "\n".join(pages)


def split_questions_from_text(text: str) -> List[str]:
    lines = [normalize_text(line) for line in text.splitlines()]
    filtered = []
    for line in lines:
        if not line:
            continue
        if line.endswith("?") or re.match(r"^\d+[\).\s]", line):
            filtered.append(line)
    return filtered


def parse_questionnaire(filename: str, file_bytes: bytes) -> List[str]:
    ext = Path(filename).suffix.lower()
    if ext == ".csv":
        return _parse_questionnaire_csv(file_bytes)
    if ext in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _parse_questionnaire_xlsx(file_bytes)
    if ext == ".pdf":
        return split_questions_from_text(parse_pdf_text(file_bytes))
    if ext in {".txt", ".md"}:
        return split_questions_from_text(file_bytes.decode("utf-8", errors="ignore"))
    raise ValueError("Unsupported questionnaire format. Upload CSV, XLSX, PDF, TXT, or MD.")


def parse_reference_doc(filename: str, file_bytes: bytes) -> str:
    ext = Path(filename).suffix.lower()
    if ext in {".txt", ".md", ".csv"}:
        return file_bytes.decode("utf-8", errors="ignore")
    if ext in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        values = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                row_items = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                if row_items:
                    values.append(" | ".join(row_items))
        return "\n".join(values)
    if ext == ".pdf":
        return parse_pdf_text(file_bytes)
    raise ValueError("Unsupported reference format. Upload TXT, MD, CSV, XLSX, or PDF.")


def _parse_questionnaire_csv(file_bytes: bytes) -> List[str]:
    decoded = file_bytes.decode("utf-8-sig", errors="ignore")
    reader = csv.DictReader(io.StringIO(decoded))
    rows = list(reader)
    if not rows:
        return []
    fieldnames = [field.strip() for field in (reader.fieldnames or []) if field]
    field_map = {field.lower(): field for field in fieldnames}
    question_field = None
    for hint in QUESTION_HINTS:
        if hint in field_map:
            question_field = field_map[hint]
            break
    if question_field is None:
        question_field = fieldnames[0]
    questions = [normalize_text(row.get(question_field, "")) for row in rows]
    return [q for q in questions if q]


def _parse_questionnaire_xlsx(file_bytes: bytes) -> List[str]:
    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    question_col = 0
    for idx, col_name in enumerate(header):
        if col_name.lower() in QUESTION_HINTS:
            question_col = idx
            break

    questions = []
    for row in rows[1:]:
        if question_col < len(row) and row[question_col] is not None:
            text = normalize_text(str(row[question_col]))
            if text:
                questions.append(text)
    return questions


def chunk_text(text: str, max_chars: int = 600, overlap: int = 120) -> List[str]:
    clean = normalize_text(text)
    if not clean:
        return []
    chunks = []
    start = 0
    while start < len(clean):
        end = min(start + max_chars, len(clean))
        chunk = clean[start:end]
        chunks.append(chunk)
        if end == len(clean):
            break
        start = max(0, end - overlap)
    return chunks
